import pandas as pd
import os
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, colors, Font, Border, Side, PatternFill
"""
    Функция обрабатывает данные о сотрудниках и их компетенциях, объединяет их в одну таблицу и записывает результаты в Excel-файл.
    Она выполняет несколько шагов: загрузка файлов, очистка данных, объединение таблиц, сортировка, создание сводной таблицы,
    форматирование и добавление гиперссылок в результирующий Excel-документ.

    Шаги выполнения:

    1. **Загрузка данных**:
       - Загрузка файлов Excel с информацией о сотрудниках и их компетенциях по заданным путям file_path_employees и file_path_comp.
       - Используется движок 'openpyxl' для работы с Excel-файлами.

    2. **Очистка и объединение данных**:
       - Удаление дубликатов в данных о компетенциях.
       - Объединение данных о сотрудниках и компетенциях по табельному номеру (поле 'Табельный номер').
       - Переименование некоторых столбцов для стандартизации названий.

    3. **Обработка столбцов**:
       - Добавление столбцов 'Фото' и 'Анкета' на основе путей к файлам, после чего исходные пути удаляются.
       - Восстановление столбцов 'Образование', 'Работа', 'Состояние' и 'Вид занятости' на свои позиции в таблице.

    4. **Сортировка данных**:
       - Сортировка сотрудников по столбцам 'Вышестоящее подразделение', 'Подразделение' и 'ФИО' с применением кастомной функции `custom_sort`.
       - Специальная сортировка для сотрудников в отпусках и из подразделений "Без подразделения" или "Магазины у дома".

    5. **Добавление разделительных строк**:
       - Создание пустых строк в таблице для визуального отделения данных по подразделениям и вышестоящим подразделениям.

    6. **Запись данных в Excel**:
       - Сортированные и обработанные данные записываются в Excel-файл.
       - Формируется сводная таблица по количеству сотрудников в каждом подразделении с добавлением строки "Итого сотрудников".

    7. **Форматирование**:
       - Устанавливается ширина столбцов для удобного отображения данных.
       - Выделение заголовков столбцов и строк, связанных с подразделениями, цветом и границами.
       - Применяются гиперссылки для путей к файлам 'Фото' и 'Анкета', если файлы существуют.
       - Применение выравнивания текста в ячейках в зависимости от содержимого.

    8. **Группировка строк**:
       - Группировка строк по вышестоящим подразделениям и подразделениям для удобства просмотра.

    9. **Финальные шаги**:
       - Удаление ненужных столбцов и закрепление первой строки таблицы.
       - Установка фильтров для первой строки.
       - Попытка сохранения файла с обработанными данными.
       - Обработка исключений при невозможности доступа к файлу (например, если файл открыт).

    10. **Результат**:
        - Данные сохраняются в указанном Excel-файле. Файл включает основную таблицу сотрудников с форматированием и сводную таблицу с численностью сотрудников по подразделениям.
    """

def competence():

    file_path_comp = os.sep * 2 + os.path.join("tg-storage01", "Служба персонала", "Общие", "Отдел аналитики", "Выгрузки. Действующие сотрудники", "Компетенции",
                                                  "Оценки компетенций сотрудников - действующие сотрудники (XLSX).xlsx")
    file_path_employees = os.sep * 2 + os.path.join("tg-storage01", "Служба персонала", "Общие", "Кадровый учет", "Действующие сотрудники.xlsx")

    # загрузка файлов
    df_emp = pd.read_excel(file_path_employees, engine='openpyxl')
    df_comp = pd.read_excel(file_path_comp, engine='openpyxl')

    # удаление повторяющихся строк
    df_comp = df_comp.drop_duplicates()
    merged_df = pd.merge(df_emp, df_comp, left_on='Табельный номер (с префиксами)', right_on='Табельный номер', how='left')
    merged_df = merged_df.drop(['Табельный номер', 'Сотрудник'], axis=1)

    merged_df = merged_df.rename(columns={'Табельный номер (с префиксами)': 'Таб. номер'})
    merged_df = merged_df.rename(columns={'Адрес по прописке представление': 'Адрес по прописке'})
    merged_df = merged_df.rename(columns={'Адрес места проживания представление': 'Адрес места проживания'})

    edu = merged_df['Образование']
    work = merged_df['Работа']
    state = merged_df['Состояние']
    type_work = merged_df['Вид занятости']
    merged_df.insert(7, 'Фото', merged_df['file_path_x'])
    merged_df.insert(17, 'Анкета', merged_df['file_path_y'])
    merged_df = merged_df.drop(['file_path_x', 'file_path_y', 'Образование', 'Работа', 'Состояние', 'Вид занятости'], axis=1)
    merged_df.insert(8, 'Образование', edu)
    merged_df.insert(9, 'Работа', work)
    merged_df.insert(21, 'Состояние', state)
    merged_df.insert(22, 'Вид занятости', type_work)


    # Сортировка данных
    # merged_df = merged_df.sort_values(by=['Вышестоящее подразделение', 'Подразделение', 'ФИО'], ascending=True)
    merged_df = merged_df.sort_values(['Вышестоящее подразделение', 'Подразделение', 'ФИО'])
    merged_df = merged_df.reset_index(drop=True)


    # Функция для определения порядка сортировки в зависимости от значения в столбце "Состояние"
    def custom_sort(row):
        if row['Состояние'] in ['Отпуск по уходу за ребенком', 'Отпуск по беременности и родам']:
            return 9999
        elif row['Вышестоящее подразделение'] in ['Без подразделения']:
            return 0
        elif row['Вышестоящее подразделение'] in ['Магазины у дома']:
            return 8888
        else:
            return row.name


    # Применение функции для сортировки
    merged_df['sort_key'] = merged_df.apply(custom_sort, axis=1)
    merged_df = merged_df.sort_values(by=['sort_key'])

    # Удаление временного столбца
    merged_df = merged_df.drop(columns=['sort_key'])
    merged_df = merged_df.reset_index(drop=True)


    # Создаем пустой список для хранения строк для объединения
    rows_to_concat = []

    # Проходим по строкам merged_df
    for i in range(len(merged_df)):
        if (i == 0) or (merged_df.loc[i - 1, 'Вышестоящее подразделение'] != merged_df.loc[i, 'Вышестоящее подразделение']):
            # Создаем пустую строку и заполняем ее значениями None
            empty_row = pd.DataFrame([[None] * len(merged_df.columns)], columns=merged_df.columns)
            # Устанавливаем значение в столбце 'Таб. номер' равным значению 'Вышестоящее подразделение' текущей строки
            empty_row.loc[0, 'Таб. номер'] = merged_df.loc[i, 'Вышестоящее подразделение']
            # Добавляем пустую строку в список rows_to_concat
            rows_to_concat.append(empty_row)
        if (i == 0) or (merged_df.loc[i - 1, 'Подразделение'] != merged_df.loc[i, 'Подразделение']) or (merged_df.loc[i - 1, 'Вышестоящее подразделение'] != merged_df.loc[i, 'Вышестоящее подразделение']):
            # Создаем пустую строку и заполняем ее значениями None
            empty_row = pd.DataFrame([[None] * len(merged_df.columns)], columns=merged_df.columns)
            # Устанавливаем значение в столбце 'Таб. номер' равным значению 'Вышестоящее подразделение' текущей строки
            empty_row.loc[0, 'ФИО'] = merged_df.loc[i, 'Подразделение']
            # Добавляем пустую строку в список rows_to_concat
            rows_to_concat.append(empty_row)

        # Добавляем текущую строку из merged_df в список rows_to_concat
        rows_to_concat.append(merged_df.iloc[[i]])
    # Создаем DataFrame из списка строк для объединения
    result_df = pd.concat(rows_to_concat, ignore_index=True)



    merged_df = result_df

    # merged_df.to_excel(file_path_employees, index=False)
    writer = pd.ExcelWriter(file_path_employees, engine='openpyxl')
    merged_df.to_excel(writer, index=False)
    # Сводная таблица по кол-ву сотрудников
    # Подсчитываем количество значений по столбцу "Вышестоящее подразделение"
    count_values = merged_df['Вышестоящее подразделение'].value_counts()

    # Общее количество строк
    total_rows = len(merged_df)

    # Создаем DataFrame с результатами
    summary_df = pd.DataFrame(count_values).reset_index()
    summary_df.columns = ['Вышестоящее подразделение', 'Количество']

    # Добавляем строку "Итого сотрудников"
    total_employees = summary_df['Количество'].sum()
    summary_df.loc[len(summary_df)] = ['Итого сотрудников', total_employees]

    # Сохраняем результат на лист "Свод" в файле file_path_employees
    # with pd.ExcelWriter(file_path_employees, engine='openpyxl', mode='a') as writer:
    summary_df.to_excel(writer, sheet_name='Свод', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Свод']
    worksheet.column_dimensions['A'].width = 55
    worksheet.column_dimensions['B'].width = 15

    # задаём ширину столбцов
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    # изменяем ширину колонки
    worksheet.column_dimensions['A'].width = 13
    worksheet.column_dimensions['B'].width = 37
    worksheet.column_dimensions['C'].width = 9
    worksheet.column_dimensions['D'].width = 35
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 50
    worksheet.column_dimensions['H'].width = 50
    worksheet.column_dimensions['I'].width = 50
    worksheet.column_dimensions['J'].width = 50
    worksheet.column_dimensions['K'].width = 50
    worksheet.column_dimensions['L'].width = 15
    worksheet.column_dimensions['M'].width = 15
    worksheet.column_dimensions['N'].width = 21
    worksheet.column_dimensions['O'].width = 30
    worksheet.column_dimensions['P'].width = 50
    worksheet.column_dimensions['Q'].width = 20
    worksheet.column_dimensions['R'].width = 30
    worksheet.column_dimensions['S'].width = 30
    worksheet.column_dimensions['T'].width = 30
    worksheet.column_dimensions['U'].width = 30

    # Проходим по каждой строке данных
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if (len(row[0]) == 0) and (len(row[3]) != 0):
            header_row = worksheet[row_num]
            fill = PatternFill(start_color="fbebdb", end_color="fbebdb", fill_type="solid")
            for cell in header_row:
                cell.fill = fill
            #     violet
        elif (len(row[0]) == 0) and (len(row[2]) != 0):
            header_row = worksheet[row_num]
            fill = PatternFill(start_color="fcdbbd", end_color="fcdbbd", fill_type="solid")
            for cell in header_row:
                cell.fill = fill
    #             pink
    group_started = False
    start_row_num = 0

    # Проходим по каждой строке данных
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):

        if (len(row[0]) == 0) and (len(row[3]) != 0):

            if not group_started:
                group_started = True
                start_row_num = row_num

        else:
            if group_started:
                group_started = False
                end_row_num = row_num
                for i in range(start_row_num, end_row_num):
                    worksheet.row_dimensions[i].outline_level = 1
                start_row_num = row_num

    # Группируем оставшиеся строки после последней непустой ячейки
    if group_started:
        end_row_num = row_num + 1
        # print(start_row_num, ",", end_row_num)
        for i in range(start_row_num, end_row_num):
            worksheet.row_dimensions[i].outline_level = 1

    group_started = False
    start_row_num = 0

    # Проходим по каждой строке данных
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row[0]) != 0:
            if not group_started:
                group_started = True
                start_row_num = row_num

        else:
            if group_started:
                group_started = False
                end_row_num = row_num
                for i in range(start_row_num, end_row_num):
                    worksheet.row_dimensions[i].outline_level = 2

                start_row_num = row_num + 1

    # Группируем оставшиеся строки после последней непустой ячейки
    if group_started:
        # print(row_num)
        end_row_num = row_num + 1
        # end_row_num = row_num
        # print(start_row_num, ",", end_row_num)
        for i in range(start_row_num, end_row_num):
            worksheet.row_dimensions[i].outline_level = 2

    worksheet.delete_cols(1, 2)

    file_paths = {
        'D': range(2, worksheet.max_row + 1),
        'O': range(2, worksheet.max_row + 1)
    }
    # Обход ячеек с путями к файлам и добавление гиперссылок
    for col, rows in file_paths.items():
        for row in rows:
            cell = worksheet[f'{col}{row}']
            file_path = cell.value
            if os.path.exists(file_path):
                hyperlink = '=HYPERLINK("{}")'.format(file_path)
                cell.value = hyperlink
                cell.font = Font(color=colors.BLUE, underline='single')


    for column in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'M', 'N', 'Q', 'R', 'T', 'O', 'P', 'L', 'S', 'U']:
        for cell in worksheet[column]:
            if cell.value: # выравниваем ячейки, только имеющие значения
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for column in ['A']:
        for cell in worksheet[column]:
            if cell.value: # выравниваем ячейки, только имеющие значения
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for column in ['G']:
        for cell in worksheet[column]:
            if cell.value:  # выравниваем ячейки, только имеющие значения
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for column in ['B', 'E']:
        for cell in worksheet[column]:
            if cell.value:  # Выделение только ячеек с данными
                cell.font = Font(bold=True)


    # ну и выравним
    for cell in worksheet[1]:
        if cell.value: # выравниваем ячейки, только имеющие значения
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Устанавливаем стиль границы для всех ячеек
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border_style

    # Устанавливаем цвет заливки для строки заголовков (первая строка)
    header_row = worksheet[1]
    fill = PatternFill(start_color="fdcc9e", end_color="fdcc9e", fill_type="solid")
    for cell in header_row:
        cell.fill = fill

    # writer._save()

    # Закрепляем первую строку
    worksheet.freeze_panes = 'A2'
    # Устанавливаем фильтр по первой строке
    worksheet.auto_filter.ref = worksheet.dimensions

    # Проверка, открыт ли файл
    try:
        workbook.close()
    except FileNotFoundError:
        # Создание нового файла, если файл не существует
        workbook = Workbook()
        workbook.save(file_path_employees)

    try:
        # для записи данных в файл
        workbook.save(file_path_employees)
    except PermissionError as e:
        print(f"Ошибка доступа к файлу: {e}")

